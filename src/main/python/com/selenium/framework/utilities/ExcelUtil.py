import openpyxl


class ExcelUtil:
    def __init__(self, driver, workbook_location=None):
        super().__init__(driver)
        self.driver = driver
        self.workbook = openpyxl.load_workbook(workbook_location)
        self.sheet = self.workbook.active
        self.cell = self.sheet.cell(row=1, column=1)

    def navigate_to_cell(self, row, col):
        return self.sheet.cell(row, col)

    def get_cell_value(self, cell):
        return self.cell.value

    def set_cell_value(self, row, col, value):
        self.navigate_to_cell(row, col)
        self.cell.value = value

    def get_max_rows(self):
        return self.sheet.max_row

    def get_max_columns(self):
        return self.sheet.max_column

    def navigate_to_cell_given_location(self, location):
        pass  # to-do

    def printSheet(self):
        for i in range(1, self.sheet.max_row + 1):
            for j in range(1, self.sheet.max_column + 1):
                print(self.sheet.cell(row=i, column=j).value)

    def capture_data_in_dictionary(self):
        new_dict = {}
        for i in range(1, self.sheet.max_row + 1):
            for j in range(1, self.sheet.max_column + 1):
                new_dict[self.sheet.cell(row=1, column=j).value] = self.sheet.cell(row=i, column=j).value
        return new_dict

    def update_cell(self, search_term, col_name, new_value):
        Dict = {}

        for i in range(1, self.sheet.max_column + 1):
            if self.sheet.cell(row=1, column=i).value == col_name:
                Dict["col"] = i

        for i in range(1, self.sheet.max_row + 1):
            for j in range(1, self.sheet.max_column + 1):
                if self.sheet.cell(row=i, column=j).value == search_term:
                    Dict["row"] = i

        self.set_cell_value(Dict["row"], Dict["col"], new_value)

    def save_workbook(self, workbook_location):
        self.workbook.save(workbook_location)
